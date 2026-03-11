"""
Sprint 196: Excel Report Upgrade Tests — "Thợ Săn Chuyên Nghiệp"

Tests for app/engine/tools/excel_report_tool.py
25 tests covering legacy mode, advanced 3-sheet mode, source collection, price extraction, and formatting.
"""

import json
import os
import pytest
import time
from pathlib import Path
from unittest.mock import patch, MagicMock


# =============================================================================
# Helper: Create test products
# =============================================================================

def _sample_products():
    return [
        {
            "platform": "Google Shopping",
            "title": "Zebra ZXP7 Printhead",
            "price": "5.500.000₫",
            "extracted_price": 5500000,
            "seller": "ABC Corp",
            "rating": 4.5,
            "sold_count": 10,
            "link": "https://abc.vn/product1",
        },
        {
            "platform": "Shopee",
            "title": "Đầu in Zebra ZXP Series 7",
            "price": "6.200.000₫",
            "extracted_price": 6200000,
            "seller": "ShopXYZ",
            "rating": 4.0,
            "link": "https://shopee.vn/product2",
        },
        {
            "platform": "International",
            "title": "Zebra ZXP7 Printhead OEM",
            "price": "7.649.745₫",
            "extracted_price": 7649745,
            "price_vnd": 7649745,
            "seller": "GlobalParts",
            "link": "https://global.com/product3",
            "dealer_info": {
                "phones": ["0901234567"],
                "zalo": ["0901234567"],
                "emails": ["sales@global.com"],
                "address": "123 Main St, US",
            },
        },
    ]


# =============================================================================
# Price Extraction Tests
# =============================================================================

class TestExtractPrice:
    """Test _extract_price()."""

    def _fn(self, price_str, extracted=None):
        from app.engine.tools.excel_report_tool import _extract_price
        return _extract_price(price_str, extracted)

    def test_extracted_price_preferred(self):
        assert self._fn("anything", 5500000) == 5500000.0

    def test_vnd_format_dots(self):
        assert self._fn("5.500.000₫") == 5500000.0

    def test_vnd_format_commas(self):
        assert self._fn("5,500,000₫") == 5500000.0

    def test_empty_string(self):
        assert self._fn("") == 0.0

    def test_no_digits(self):
        assert self._fn("Liên hệ") == 0.0


class TestFormatVnd:
    """Test _format_vnd()."""

    def _fn(self, price):
        from app.engine.tools.excel_report_tool import _format_vnd
        return _format_vnd(price)

    def test_standard(self):
        result = self._fn(5500000)
        assert "5.500.000" in result
        assert "₫" in result

    def test_zero(self):
        assert self._fn(0) == "N/A"

    def test_none(self):
        assert self._fn(None) == "N/A"


# =============================================================================
# Source Collection Tests (Sprint 196)
# =============================================================================

class TestCollectSources:
    """Test _collect_sources()."""

    def _fn(self, products):
        from app.engine.tools.excel_report_tool import _collect_sources
        return _collect_sources(products)

    def test_basic_sources(self):
        products = _sample_products()
        sources = self._fn(products)
        assert len(sources) >= 2

    def test_dealer_type(self):
        products = [
            {"platform": "Web", "seller": "Dealer A", "dealer_info": {"phones": ["0901234567"], "zalo": [], "emails": [], "address": ""}},
        ]
        sources = self._fn(products)
        assert sources[0]["type"] == "dealer"

    def test_ecommerce_type(self):
        products = [
            {"platform": "Shopee", "seller": "ShopABC"},
        ]
        sources = self._fn(products)
        assert sources[0]["type"] == "e-commerce"

    def test_deduplication(self):
        products = [
            {"platform": "Shopee", "seller": "ShopA", "title": "Product 1"},
            {"platform": "Shopee", "seller": "ShopA", "title": "Product 2"},
        ]
        sources = self._fn(products)
        assert len(sources) == 1

    def test_international_type(self):
        products = [
            {"platform": "International", "seller": "GlobalCorp"},
        ]
        sources = self._fn(products)
        assert sources[0]["type"] == "international"


# =============================================================================
# Legacy Report Tests
# =============================================================================

class TestLegacyReport:
    """Test _generate_legacy_report() (xlsxwriter mode)."""

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_generates_file(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_legacy_report
        result = json.loads(_generate_legacy_report(products, "Test Report"))
        assert "file_path" in result
        assert result["total_products"] == 3
        assert Path(result["file_path"]).exists()

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_sorted_by_price(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_legacy_report
        result = json.loads(_generate_legacy_report(products, "Test"))
        assert result["min_price"] == 5500000

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_platforms_listed(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_legacy_report
        result = json.loads(_generate_legacy_report(products, "Test"))
        assert len(result["platforms"]) >= 2


# =============================================================================
# Advanced Report Tests (Sprint 196)
# =============================================================================

class TestAdvancedReport:
    """Test _generate_advanced_report() (openpyxl 3-sheet mode)."""

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_generates_3_sheets(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_advanced_report
        result = json.loads(_generate_advanced_report(products, "Advanced Report"))
        assert "sheets" in result
        assert len(result["sheets"]) == 3
        assert "Nguồn Cung Cấp" in result["sheets"]
        assert "Thông Tin Sản Phẩm" in result["sheets"]
        assert "Khuyến Nghị" in result["sheets"]

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_file_created(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_advanced_report
        result = json.loads(_generate_advanced_report(products, "Test"))
        assert Path(result["file_path"]).exists()

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_total_sources_count(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_advanced_report
        result = json.loads(_generate_advanced_report(products, "Test"))
        assert result["total_sources"] >= 2

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_min_max_prices(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_advanced_report
        result = json.loads(_generate_advanced_report(products, "Test"))
        assert result["min_price"] is not None
        assert result["max_price"] is not None
        assert result["min_price"] <= result["max_price"]

    @patch("app.engine.tools.excel_report_tool._get_reports_dir")
    def test_openpyxl_workbook_structure(self, mock_dir, tmp_path):
        mock_dir.return_value = tmp_path
        products = _sample_products()

        from app.engine.tools.excel_report_tool import _generate_advanced_report
        result = json.loads(_generate_advanced_report(products, "Test"))

        # Verify openpyxl can read it back
        from openpyxl import load_workbook
        wb = load_workbook(result["file_path"])
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames[0] == "Nguồn Cung Cấp"
        assert wb.sheetnames[1] == "Thông Tin Sản Phẩm"
        assert wb.sheetnames[2] == "Khuyến Nghị"
        wb.close()


# =============================================================================
# Tool Function Dispatch Tests
# =============================================================================

class TestToolFunctionDispatch:
    """Test tool_generate_product_report dispatches correctly."""

    @patch("app.engine.tools.excel_report_tool._generate_legacy_report")
    @patch("app.engine.tools.excel_report_tool._advanced_excel_report_enabled", return_value=False)
    def test_flag_off_uses_legacy(self, mock_flag, mock_legacy):
        mock_legacy.return_value = json.dumps({"file_path": "/tmp/test.xlsx"})

        from app.engine.tools.excel_report_tool import tool_generate_product_report
        result = tool_generate_product_report.invoke({"products_json": json.dumps([{"platform": "test", "title": "t"}])})
        mock_legacy.assert_called_once()

    @patch("app.engine.tools.excel_report_tool._generate_advanced_report")
    @patch("app.core.config.get_settings")
    def test_flag_on_uses_advanced(self, mock_settings, mock_advanced):
        mock_settings.return_value = MagicMock(enable_advanced_excel_report=True)
        mock_advanced.return_value = json.dumps({"file_path": "/tmp/test.xlsx", "sheets": []})

        from app.engine.tools.excel_report_tool import tool_generate_product_report
        result = tool_generate_product_report.invoke({"products_json": json.dumps([{"platform": "test", "title": "t"}])})
        mock_advanced.assert_called_once()

    def test_invalid_json_input(self):
        from app.engine.tools.excel_report_tool import tool_generate_product_report
        result = json.loads(tool_generate_product_report.invoke({"products_json": "not json"}))
        assert "error" in result

    def test_empty_products(self):
        from app.engine.tools.excel_report_tool import tool_generate_product_report
        result = json.loads(tool_generate_product_report.invoke({"products_json": "[]"}))
        assert "error" in result

    def test_not_array(self):
        from app.engine.tools.excel_report_tool import tool_generate_product_report
        result = json.loads(tool_generate_product_report.invoke({"products_json": '{"key": "val"}'}))
        assert "error" in result
