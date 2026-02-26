"""
Excel Report Tool — Sprint 148: "Săn Hàng" + Sprint 196: "Thợ Săn Chuyên Nghiệp"

Sprint 148: Single-sheet product comparison (XlsxWriter).
Sprint 196: Advanced 3-sheet report (openpyxl) with dealer contacts + recommendations.
Gate: enable_advanced_excel_report (False → legacy single-sheet behavior).

Output: ~/.wiii/workspace/reports/product_report_{timestamp}.xlsx
"""

import json
import logging
import re
import time
from pathlib import Path

from langchain_core.tools import tool

from app.engine.tools.registry import (
    ToolCategory,
    ToolAccess,
    get_tool_registry,
)

logger = logging.getLogger(__name__)


def _get_reports_dir() -> Path:
    """Get the reports directory, creating it if needed."""
    from app.core.config import get_settings
    settings = get_settings()
    workspace = Path(settings.workspace_root).expanduser()
    reports_dir = workspace / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    return reports_dir


@tool
def tool_generate_product_report(products_json: str, title: str = "Báo cáo so sánh sản phẩm") -> str:
    """Generate an Excel report comparing products across platforms.
    Creates a formatted .xlsx file with product data, pricing, and links.
    When advanced mode is enabled, generates 3 sheets: Sources, Products, Recommendations.

    Args:
        products_json: JSON string of product list. Each item should have:
            platform, title, price, seller, rating, sold_count, delivery, location, link.
            Optional (Sprint 196): dealer_info (dict with phones, zalo, email, address),
            product_type (part/machine/accessory/service), price_vnd (for international items).
        title: Report title (default: "Báo cáo so sánh sản phẩm")

    Returns:
        Path to the generated Excel file, or error message.
    """
    try:
        products = json.loads(products_json)
        if not isinstance(products, list):
            return json.dumps({"error": "products_json must be a JSON array"}, ensure_ascii=False)
    except json.JSONDecodeError as e:
        return json.dumps({"error": f"Invalid JSON: {str(e)[:100]}"}, ensure_ascii=False)

    if not products:
        return json.dumps({"error": "No products to report"}, ensure_ascii=False)

    # Sprint 196: Check if advanced mode is enabled
    from app.core.config import get_settings
    settings = get_settings()
    if settings.enable_advanced_excel_report:
        return _generate_advanced_report(products, title)
    else:
        return _generate_legacy_report(products, title)


def _generate_legacy_report(products: list, title: str) -> str:
    """Generate legacy single-sheet report using xlsxwriter (Sprint 148 behavior)."""
    try:
        import xlsxwriter
    except ImportError:
        return json.dumps({"error": "xlsxwriter not installed. Run: pip install xlsxwriter"}, ensure_ascii=False)

    reports_dir = _get_reports_dir()
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"product_report_{timestamp}.xlsx"
    filepath = reports_dir / filename

    try:
        workbook = xlsxwriter.Workbook(str(filepath))
        worksheet = workbook.add_worksheet("So sánh sản phẩm")

        # Formats
        header_fmt = workbook.add_format({
            "bold": True,
            "bg_color": "#FF8C00",
            "font_color": "#FFFFFF",
            "border": 1,
            "text_wrap": True,
            "valign": "vcenter",
            "align": "center",
        })
        price_fmt = workbook.add_format({
            "num_format": "#,##0 ₫",
            "border": 1,
        })
        cell_fmt = workbook.add_format({
            "border": 1,
            "text_wrap": True,
            "valign": "top",
        })
        link_fmt = workbook.add_format({
            "border": 1,
            "font_color": "#0563C1",
            "underline": True,
        })
        min_price_fmt = workbook.add_format({
            "num_format": "#,##0 ₫",
            "border": 1,
            "bg_color": "#C6EFCE",
            "font_color": "#006100",
            "bold": True,
        })
        max_price_fmt = workbook.add_format({
            "num_format": "#,##0 ₫",
            "border": 1,
            "bg_color": "#FFC7CE",
            "font_color": "#9C0006",
        })
        title_fmt = workbook.add_format({
            "bold": True,
            "font_size": 14,
        })

        # Title row
        worksheet.merge_range("A1:K1", title, title_fmt)
        worksheet.write("A2", f"Ngày tạo: {time.strftime('%d/%m/%Y %H:%M')}", workbook.add_format({"italic": True}))

        # Headers
        headers = ["STT", "Sàn", "Tên SP", "Giá (VNĐ)", "Người bán", "Đánh giá", "Lượt bán", "Vận chuyển", "Địa chỉ", "Link", "Mô tả"]
        col_widths = [5, 18, 45, 16, 25, 10, 12, 15, 20, 35, 35]
        for col, (header, width) in enumerate(zip(headers, col_widths)):
            worksheet.set_column(col, col, width)
            worksheet.write(3, col, header, header_fmt)

        prices = [_extract_price(p.get("price", ""), p.get("extracted_price")) for p in products]
        paired = list(zip(products, prices))
        paired.sort(key=lambda x: x[1] if x[1] and x[1] > 0 else float('inf'))
        products = [p for p, _ in paired]
        prices = [pr for _, pr in paired]

        valid_prices = [p for p in prices if p and p > 0]
        min_price = min(valid_prices) if valid_prices else None
        max_price = max(valid_prices) if valid_prices else None

        for i, product in enumerate(products):
            row = 4 + i
            price_val = prices[i]
            p_fmt = price_fmt
            if price_val and min_price and price_val == min_price:
                p_fmt = min_price_fmt
            elif price_val and max_price and price_val == max_price:
                p_fmt = max_price_fmt

            worksheet.write(row, 0, i + 1, cell_fmt)
            worksheet.write(row, 1, product.get("platform", ""), cell_fmt)
            worksheet.write(row, 2, product.get("title", ""), cell_fmt)
            worksheet.write(row, 3, price_val or 0, p_fmt)
            worksheet.write(row, 4, product.get("seller", "") or product.get("source", ""), cell_fmt)
            worksheet.write(row, 5, product.get("rating", "") or "", cell_fmt)
            worksheet.write(row, 6, product.get("sold_count", "") or "", cell_fmt)
            worksheet.write(row, 7, product.get("delivery", ""), cell_fmt)
            worksheet.write(row, 8, product.get("location", ""), cell_fmt)

            link = product.get("link") or product.get("url", "")
            if link:
                worksheet.write_url(row, 9, link, link_fmt, "Xem SP")
            else:
                worksheet.write(row, 9, "", cell_fmt)

            description = product.get("snippet", "") or product.get("description", "")
            worksheet.write(row, 10, description, cell_fmt)

        # Summary row
        summary_row = 4 + len(products) + 1
        summary_fmt = workbook.add_format({"bold": True, "border": 1, "bg_color": "#F2F2F2"})
        worksheet.write(summary_row, 0, "", summary_fmt)
        worksheet.write(summary_row, 1, "TỔNG KẾT", summary_fmt)
        worksheet.write(summary_row, 2, f"{len(products)} sản phẩm từ {len(set(p.get('platform', '') for p in products))} sàn", summary_fmt)
        if min_price:
            worksheet.write(summary_row, 3, min_price, min_price_fmt)
        worksheet.write(summary_row, 4, f"Giá thấp nhất: {_format_vnd(min_price)}" if min_price else "", summary_fmt)

        workbook.close()

        return json.dumps({
            "file_path": str(filepath),
            "filename": filename,
            "total_products": len(products),
            "platforms": list(set(p.get("platform", "") for p in products)),
            "min_price": min_price,
            "max_price": max_price,
        }, ensure_ascii=False)

    except Exception as e:
        logger.error("[EXCEL_REPORT] Legacy generation failed: %s", e)
        return json.dumps({"error": f"Lỗi tạo báo cáo: {str(e)[:200]}"}, ensure_ascii=False)


def _generate_advanced_report(products: list, title: str) -> str:
    """Generate advanced 3-sheet report using openpyxl (Sprint 196).

    Sheet 1: "Nguồn Cung Cấp" — dealer/source contact info
    Sheet 2: "Thông Tin Sản Phẩm" — product details with pricing
    Sheet 3: "Khuyến Nghị" — recommendations (cheapest, most reliable, comparison)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}, ensure_ascii=False)

    reports_dir = _get_reports_dir()
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"product_report_{timestamp}.xlsx"
    filepath = reports_dir / filename

    try:
        wb = Workbook()

        # Shared styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        title_font = Font(bold=True, size=14)

        # ========== Sheet 1: Nguồn Cung Cấp ==========
        ws1 = wb.active
        ws1.title = "Nguồn Cung Cấp"
        ws1.merge_cells("A1:H1")
        ws1["A1"] = title
        ws1["A1"].font = title_font
        ws1["A2"] = f"Ngày tạo: {time.strftime('%d/%m/%Y %H:%M')}"
        ws1["A2"].font = Font(italic=True)

        source_headers = ["STT", "Nguồn", "Loại", "SĐT", "Zalo", "Email", "Địa chỉ", "Website"]
        source_widths = [5, 30, 15, 18, 18, 25, 35, 35]
        for col_idx, (header, width) in enumerate(zip(source_headers, source_widths), 1):
            cell = ws1.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws1.column_dimensions[cell.column_letter].width = width

        # Collect unique sources
        sources = _collect_sources(products)
        for i, src in enumerate(sources):
            row = 5 + i
            values = [
                i + 1,
                src.get("name", ""),
                src.get("type", ""),
                ", ".join(src.get("phones", [])),
                ", ".join(src.get("zalo", [])),
                ", ".join(src.get("emails", [])),
                src.get("address", ""),
                src.get("url", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws1.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = cell_alignment

        # ========== Sheet 2: Thông Tin Sản Phẩm ==========
        ws2 = wb.create_sheet("Thông Tin Sản Phẩm")
        ws2.merge_cells("A1:J1")
        ws2["A1"] = title
        ws2["A1"].font = title_font

        product_headers = ["STT", "Tên SP", "Giá (VNĐ)", "Sàn/Nguồn", "Link", "Loại SP", "Tình trạng", "Người bán", "Đánh giá", "Mô tả"]
        product_widths = [5, 45, 18, 20, 35, 12, 12, 25, 10, 35]
        for col_idx, (header, width) in enumerate(zip(product_headers, product_widths), 1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws2.column_dimensions[cell.column_letter].width = width

        # Sort products by price
        prices = [_extract_price(p.get("price", ""), p.get("extracted_price", p.get("price_vnd"))) for p in products]
        paired = list(zip(products, prices))
        paired.sort(key=lambda x: x[1] if x[1] and x[1] > 0 else float('inf'))
        products_sorted = [p for p, _ in paired]
        prices_sorted = [pr for _, pr in paired]

        valid_prices = [p for p in prices_sorted if p and p > 0]
        min_price = min(valid_prices) if valid_prices else None
        max_price = max(valid_prices) if valid_prices else None

        vnd_format = '#,##0'
        for i, (product, price_val) in enumerate(zip(products_sorted, prices_sorted)):
            row = 4 + i
            ws2.cell(row=row, column=1, value=i + 1).border = thin_border

            ws2.cell(row=row, column=2, value=product.get("title", "")).border = thin_border

            price_cell = ws2.cell(row=row, column=3, value=price_val or 0)
            price_cell.number_format = vnd_format
            price_cell.border = thin_border
            if price_val and min_price and price_val == min_price:
                price_cell.fill = green_fill
                price_cell.font = green_font
            elif price_val and max_price and price_val == max_price:
                price_cell.fill = red_fill

            ws2.cell(row=row, column=4, value=product.get("platform", "")).border = thin_border

            link = product.get("link") or product.get("url", "")
            link_cell = ws2.cell(row=row, column=5, value=link)
            link_cell.border = thin_border
            if link:
                link_cell.hyperlink = link
                link_cell.font = Font(color="0563C1", underline="single")

            ws2.cell(row=row, column=6, value=product.get("product_type", "")).border = thin_border
            ws2.cell(row=row, column=7, value=product.get("condition", "")).border = thin_border
            ws2.cell(row=row, column=8, value=product.get("seller", "") or product.get("source", "")).border = thin_border
            ws2.cell(row=row, column=9, value=product.get("rating", "") or "").border = thin_border

            desc = product.get("snippet", "") or product.get("description", "")
            ws2.cell(row=row, column=10, value=desc).border = thin_border

            for col in range(1, 11):
                ws2.cell(row=row, column=col).alignment = cell_alignment

        # ========== Sheet 3: Khuyến Nghị ==========
        ws3 = wb.create_sheet("Khuyến Nghị")
        ws3.merge_cells("A1:D1")
        ws3["A1"] = "Khuyến Nghị Mua Hàng"
        ws3["A1"].font = title_font
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 45
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 35

        rec_headers = ["Hạng mục", "Sản phẩm", "Giá (VNĐ)", "Nguồn"]
        for col_idx, header in enumerate(rec_headers, 1):
            cell = ws3.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Top 3 cheapest
        row_offset = 4
        ws3.cell(row=row_offset, column=1, value="GIÁ RẺ NHẤT").font = Font(bold=True, size=12)
        cheapest = [(p, pr) for p, pr in zip(products_sorted, prices_sorted) if pr and pr > 0][:3]
        for i, (prod, price) in enumerate(cheapest):
            r = row_offset + 1 + i
            ws3.cell(row=r, column=1, value=f"Top {i + 1}").border = thin_border
            ws3.cell(row=r, column=2, value=prod.get("title", "")[:60]).border = thin_border
            p_cell = ws3.cell(row=r, column=3, value=price)
            p_cell.number_format = vnd_format
            p_cell.border = thin_border
            if i == 0:
                p_cell.fill = green_fill
                p_cell.font = green_font
            ws3.cell(row=r, column=4, value=prod.get("platform", "")).border = thin_border

        # Top 3 most reliable (by rating)
        row_offset = row_offset + len(cheapest) + 2
        ws3.cell(row=row_offset, column=1, value="UY TÍN NHẤT").font = Font(bold=True, size=12)
        rated = sorted(
            [(p, pr) for p, pr in zip(products_sorted, prices_sorted)],
            key=lambda x: float(x[0].get("rating", 0) or 0),
            reverse=True,
        )[:3]
        for i, (prod, price) in enumerate(rated):
            r = row_offset + 1 + i
            ws3.cell(row=r, column=1, value=f"Top {i + 1} (⭐{prod.get('rating', 'N/A')})").border = thin_border
            ws3.cell(row=r, column=2, value=prod.get("title", "")[:60]).border = thin_border
            p_cell = ws3.cell(row=r, column=3, value=price or 0)
            p_cell.number_format = vnd_format
            p_cell.border = thin_border
            ws3.cell(row=r, column=4, value=prod.get("platform", "")).border = thin_border

        # Domestic vs International comparison
        row_offset = row_offset + len(rated) + 2
        ws3.cell(row=row_offset, column=1, value="NỘI ĐỊA vs QUỐC TẾ").font = Font(bold=True, size=12)
        domestic = [p for p in products_sorted if p.get("platform", "").lower() not in ("international", "global", "worldwide")]
        international = [p for p in products_sorted if p.get("platform", "").lower() in ("international", "global", "worldwide") or p.get("price_vnd")]

        domestic_prices = [_extract_price(p.get("price", ""), p.get("extracted_price")) for p in domestic]
        intl_prices = [p.get("price_vnd", 0) or _extract_price(p.get("price", ""), p.get("extracted_price")) for p in international]

        domestic_min = min([p for p in domestic_prices if p and p > 0], default=0)
        intl_min = min([p for p in intl_prices if p and p > 0], default=0)

        r = row_offset + 1
        ws3.cell(row=r, column=1, value="Giá nội địa thấp nhất").border = thin_border
        p_cell = ws3.cell(row=r, column=3, value=domestic_min or 0)
        p_cell.number_format = vnd_format
        p_cell.border = thin_border
        ws3.cell(row=r, column=4, value=f"{len(domestic)} nguồn nội địa").border = thin_border

        r += 1
        ws3.cell(row=r, column=1, value="Giá quốc tế thấp nhất").border = thin_border
        p_cell = ws3.cell(row=r, column=3, value=intl_min or 0)
        p_cell.number_format = vnd_format
        p_cell.border = thin_border
        ws3.cell(row=r, column=4, value=f"{len(international)} nguồn quốc tế").border = thin_border

        # Summary
        r += 2
        ws3.cell(row=r, column=1, value="TỔNG KẾT").font = Font(bold=True, size=12)
        r += 1
        ws3.cell(row=r, column=1, value="Tổng sản phẩm").border = thin_border
        ws3.cell(row=r, column=2, value=len(products)).border = thin_border
        r += 1
        ws3.cell(row=r, column=1, value="Tổng nguồn").border = thin_border
        ws3.cell(row=r, column=2, value=len(set(p.get("platform", "") for p in products))).border = thin_border
        r += 1
        ws3.cell(row=r, column=1, value="Giá thấp nhất toàn bộ").border = thin_border
        p_cell = ws3.cell(row=r, column=2, value=min_price or 0)
        p_cell.number_format = vnd_format
        p_cell.border = thin_border
        p_cell.fill = green_fill
        p_cell.font = green_font

        wb.save(str(filepath))

        return json.dumps({
            "file_path": str(filepath),
            "filename": filename,
            "total_products": len(products),
            "total_sources": len(sources),
            "sheets": ["Nguồn Cung Cấp", "Thông Tin Sản Phẩm", "Khuyến Nghị"],
            "platforms": list(set(p.get("platform", "") for p in products)),
            "min_price": min_price,
            "max_price": max_price,
        }, ensure_ascii=False)

    except Exception as e:
        logger.error("[EXCEL_REPORT] Advanced generation failed: %s", e)
        return json.dumps({"error": f"Lỗi tạo báo cáo nâng cao: {str(e)[:200]}"}, ensure_ascii=False)


def _collect_sources(products: list) -> list:
    """Collect unique sources from products for the Sources sheet."""
    seen = set()
    sources = []
    for p in products:
        platform = p.get("platform", "")
        seller = p.get("seller", "") or p.get("source", "")
        key = f"{platform}:{seller}"
        if key in seen:
            continue
        seen.add(key)

        dealer = p.get("dealer_info", {}) or {}
        source_type = "e-commerce"
        if dealer:
            source_type = "dealer"
        elif platform.lower() in ("international", "global", "worldwide"):
            source_type = "international"
        elif "facebook" in platform.lower():
            source_type = "social"

        sources.append({
            "name": seller or platform,
            "type": source_type,
            "phones": dealer.get("phones", []) if dealer else [],
            "zalo": dealer.get("zalo", []) if dealer else [],
            "emails": dealer.get("emails", []) if dealer else [],
            "address": dealer.get("address", "") if dealer else "",
            "url": p.get("link", "") or p.get("url", ""),
        })
    return sources


def _extract_price(price_str, extracted_price=None) -> float:
    """Extract numeric price from various formats."""
    if extracted_price and isinstance(extracted_price, (int, float)):
        return float(extracted_price)
    if not price_str:
        return 0.0
    price_str = str(price_str)
    cleaned = re.sub(r'[^\d.,]', '', price_str)
    # Vietnamese format: both . and , are thousands separators (VND has no decimals)
    cleaned = cleaned.replace('.', '').replace(',', '')
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def _format_vnd(price: float) -> str:
    """Format price as Vietnamese Dong."""
    if not price:
        return "N/A"
    return f"{price:,.0f}₫".replace(",", ".")


# =============================================================================
# Registration
# =============================================================================

def init_excel_report_tool():
    """Register Excel report tool with the global registry."""
    registry = get_tool_registry()
    registry.register(tool_generate_product_report, ToolCategory.PRODUCT_SEARCH, ToolAccess.WRITE)
    logger.info("Excel report tool registered")
